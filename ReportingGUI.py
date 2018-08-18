import csv
import datetime
import glob
import sys
from mmap import mmap, ACCESS_READ
from tkinter import Tk
from tkinter.filedialog import askopenfilename

from openpyxl import workbook, load_workbook
import xlrd
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QMessageBox


class App(QWidget):

    def __init__(self):
        super().__init__()
        self.title = 'Report Generation'
        self.left = 100
        self.top = 100
        self.width = 300
        self.height = 70
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        description = "Please select which operation you'd like to perform."
        desc_label = QLabel(description, self)
        desc_label.move(26, 10)

        call_log_button = QPushButton('Call Logs', self)
        ops_report_button = QPushButton('Operations Report', self)
        call_log_button.clicked.connect(self.call_click)
        ops_report_button.clicked.connect(self.ops_click)
        call_log_button.move(50, 40)
        ops_report_button.move(150, 40)

        self.show()

    @pyqtSlot()
    def ops_click(self):
        try:
            self.ops_summary()
            QMessageBox.error(self, 'Finished', 'Operations Report Created.')
        except:
            QMessageBox.error(self, 'Error', "Looks like something went wrong.\n"
                                             "Check to make sure the file being written to isn't open.")

    @pyqtSlot()
    def call_click(self):
        try:
            self.call_logs()
            QMessageBox.about(self, 'Finished', 'Call Report Created.')
        except:
            QMessageBox.about(self, 'Error', "Looks like something went wrong.\n"
                                             "Check to make sure the file being written to isn't open.")

    @staticmethod
    def call_logs():

        now = datetime.datetime.now().date()

        newFile = csv.writer(open("../../files/combinedList" + str(now) + ".csv", "w+", newline=''))
        fullSheet = []
        communityReport = []
        personReport = []
        person = {}
        personCalls = {}
        personNotes = {}
        Tk().withdraw()
        thisfile = askopenfilename()

        with open(thisfile, 'rb') as currentfile:
            wrkbook = load_workbook(currentfile, read_only=True)
            ws = wrkbook.get_active_sheet()
            print(ws)

            if ws:
                communitycalls = {}
                communitynotes = {}
                communities = {}
                personcommunity = {}
                for row in ws.rows:
                    name = ""
                    contact = ""
                    community = ""
                    for cell in row:
                        if cell.row != 1:
                            if cell.column == 1:
                                community = cell.value
                                if community not in communities:
                                    communities[community] = 1
                                    communitynotes[community] = 0
                                    communitycalls[community] = 0
                                else:
                                    communities[community] += 1
                            elif cell.column == 2:
                                if cell.value:
                                    name = cell.value
                                else:
                                    name = "NULL"
                                if name not in person:
                                    person[name] = 1
                                    personCalls[name] = 0
                                    personNotes[name] = 0
                                else:
                                    person[name] += 1
                                if community not in personcommunity:
                                    personcommunity[community] = {}
                                if name not in personcommunity[community]:
                                    personcommunity[community][name] = {"call": 0, "note": 0}
                            elif cell.column == 6:
                                contact = cell.value
                                if contact == "note":
                                    personNotes[name] += 1
                                    communitynotes[community] += 1
                                if contact == "call":
                                    personCalls[name] += 1
                                    communitycalls[community] += 1
                                personcommunity[community][name][contact] += 1
                            else:
                                pass
                total = 0
                totalcalls = 0
                totalnotes = 0
                fullSheet.append(['']+['Total Notes']+['Total Calls']+['Notes+Calls'])
                for key, value in person.items():
                    total = total + value
                for key, value in personCalls.items():
                    totalcalls = totalcalls + value
                for key, value in personNotes.items():
                    totalnotes = totalnotes + value
                fullSheet.append(['Total'] + [totalnotes] + [totalcalls] + [total])
                fullSheet.append([''])
                fullSheet.append(['Community']+['Person']+['Calls']+['Notes']+['Notes+Calls'])

                for key, value in personcommunity.items():
                    fullSheet.append([key])
                    thislist = []
                    for name, nums in value.items():
                        for each, vals in nums.items():
                            if each == 'call':
                                thislist.append([name]+[str(vals)]+[str(nums['note'])]+[vals+nums['note']])
                                print(str(thislist))
                    thislist = sorted(thislist, key=lambda z: z[3], reverse=True)
                    print(thislist)
                    for each in thislist:
                        fullSheet.append(each)
                    fullSheet.append([''])

                for key, value in communities.items():
                    print(key)
                    communityReport.append([key] + [communitynotes[key]] + [communitycalls[key]] +
                                           [communitycalls[key]+communitynotes[key]])

                for key, value in person.items():
                    print(key)
                    personReport.append([key] + [str(personNotes[key])] + [str(personCalls[key])] + [str(person[key])])

                for line in fullSheet:
                    newFile.writerow(line)

                newFile.writerow('')
                newFile.writerow(['Community'] + ['Total Notes'] + ['Total Calls'] + ['Notes + Calls'])
                communityReport = sorted(communityReport, key=lambda z: z[3], reverse=True)
                for line in communityReport:
                    newFile.writerow(line)

                for each in personReport:
                    each[3] = float(each[3])
                newFile.writerow('')
                newFile.writerow(['Person'] + ['Total Notes'] + ['Total Calls'] + ['Notes + Calls'])
                personReport = sorted(personReport, key=lambda z: z[3], reverse=True)
                for line in personReport:
                    newFile.writerow(line)

                newFile.writerow(["Person/Community"] + ["Notes"] + ["Calls"])
                for key, value in personcommunity.items():
                    if key[-4:] == "call":
                        newFile.writerow([key[:-4]])


@staticmethod
def ops_summary():
    now = datetime.datetime.now().date()
    date_range = ""
    communities = {}

    newFile = csv.writer(open("../../files/ops_summary" + str(now) + ".csv", "w+", newline=''))
    for xlsFile in glob.glob("../../files/Operations_Summary_*.xls"):
        with open(xlsFile, 'rb') as currentfile:
            workbook = xlrd.open_workbook(file_contents=mmap(currentfile.fileno(), 0, access=ACCESS_READ))

            for s in workbook.sheets():
                if date_range == "":
                    date_range = s.cell(3, 0).value
                if s.cell(3, 0).value == date_range:
                    communityname = s.cell(1, 0).value
                    for row in range(s.nrows):
                        if s.cell(row, 0).value == "Agendas" and s.cell(row, 1).value == "Published":
                            agendas = s.cell(row, 3).value
                            print("agenda")
                        elif s.cell(row, 0).value == "Management Reports" and s.cell(row, 1).value == "Published":
                            management = s.cell(row, 3).value
                            print("management")
                        elif s.cell(row, 0).value == "Meeting Minutes" and s.cell(row, 1).value == "Published":
                            minutes = s.cell(row, 3).value
                            print("minutes")
                        elif s.cell(row, 0).value == "Architectural Control" and s.cell(row, 1).value == "Total":
                            arch_cont = s.cell(row, 3).value
                            print("arch")
                        elif s.cell(row, 0).value == "Violations" and s.cell(row, 1).value == "Open" and \
                                s.cell(row, 2).value == "Created":
                            violations = s.cell(row, 3).value
                            print("viol")
                        elif s.cell(row, 0).value == "Maintenance" and s.cell(row, 1).value == "Open" and \
                                s.cell(row, 2).value == "Created":
                            maint = s.cell(row, 3).value
                            print("maint")
                        elif s.cell(row, 0).value == "Broadcast Messages" and s.cell(row, 1).value == "Sent":
                            broadcast = s.cell(row, 3).value
                            print("b")
                        elif s.cell(row, 0).value == "Newsletters" and s.cell(row, 1).value == "Published":
                            newsletters = s.cell(row, 3).value
                            print("n")
                        elif s.cell(row, 0).value == "Calls" and s.cell(row, 1).value == "Calls":
                            calls = s.cell(row, 3).value
                            print("c")
                    communities[communityname] = [agendas, management, minutes, arch_cont, violations, maint,
                                                  broadcast, newsletters, calls]
                    print("Added " + communityname)

    full_agendas = 0
    full_management = 0
    full_minutes = 0
    full_arch_count = 0
    full_violations = 0
    full_maint = 0
    full_broadcast = 0
    full_newsletters = 0
    full_calls = 0
    for name, community in communities.items():
        full_agendas += float(community[0])
        full_management += float(community[1])
        full_minutes += int(community[2])
        full_arch_count += int(community[3])
        full_violations += int(community[4])
        full_maint += int(community[5])
        full_broadcast += int(community[6])
        full_newsletters += int(community[7])
        full_calls += int(community[8])

    newFile.writerow([date_range])
    newFile.writerow('')
    newFile.writerow(['Summary'] + [''] + ['Grand totals'])
    newFile.writerow('')
    newFile.writerow('')
    newFile.writerow(['Agendas'] + [''] + [full_agendas])
    newFile.writerow(['Management reports'] + [''] + [full_management])
    newFile.writerow(['Meeting Minutes'] + [''] + [full_minutes])
    newFile.writerow(['Architectural control'] + [''] + [full_arch_count])
    newFile.writerow(['Violations created'] + [''] + [full_violations])
    newFile.writerow(['Maintenance Created'] + [''] + [full_maint])
    newFile.writerow(['Broadcast messages sent'] + [''] + [full_broadcast])
    newFile.writerow(['Newsletters'] + [''] + [full_newsletters])
    newFile.writerow(['Calls'] + [''] + [full_calls])
    newFile.writerow('')
    newFile.writerow('')
    newFile.writerow('')
    x = 0
    for community, values in communities.items():
        newFile.writerow([community])
        newFile.writerow("")
        newFile.writerow(["Agendas"] + [""] + [values[0]])
        newFile.writerow(["Management reports"] + [""] + [values[1]])
        newFile.writerow(["Meeting Minutes"] + [""] + [values[2]])
        newFile.writerow(["Architectural control"] + [""] + [values[3]])
        newFile.writerow(["Violations created"] + [""] + [values[4]])
        newFile.writerow(["Maintenance Created"] + [""] + [values[5]])
        newFile.writerow(["Broadcast messages sent"] + [""] + [values[6]])
        newFile.writerow(["Newsletters"] + [""] + [values[7]])
        newFile.writerow(["Calls"] + [""] + [values[8]])
        newFile.writerow("")
        newFile.writerow("")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())
